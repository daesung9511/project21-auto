import os
import random
import time
import psutil
from shutil import copyfile
from datetime import datetime
from pathlib import Path

from os import listdir
from os.path import isfile, join

from selenium.webdriver.android.webdriver import WebDriver
from selenium import webdriver
import chromedriver_binary  # Adds chromedriver binary to path
import datetime
from pathlib import Path
from dataclasses import dataclass

from selenium.webdriver.remote.webelement import WebElement

from config import CHROME_USER_DATA_PATH, CHROME_PROFILE_NAME, CHROME_GFA_PROFILE_NAME, KEY_SHEET_FILE_PATH, \
    RAW_FILE_PATH, RD_FILE

from openpyxl import Workbook, worksheet, load_workbook
import fnmatch

DEFAULT_TIMEOUT_DELAY = 5


@dataclass
class UserConfig:
    user_data_path: str
    profile_name: str
    gfa_profile_name: str
    download_path: str


class Utils:

    @staticmethod
    def get_chrome_driver() -> WebDriver:

        options = webdriver.ChromeOptions()
        # https://www.python2.net/questions-80772.htm
        options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/87.0.4280.141 Safari/537.36")
        options.add_experimental_option("detach", True)
        user_config = Utils._get_config()
        prefs = {
            "profile.default_content_settings.popups": 0,
            f'download.default_directory': user_config.download_path,
            "directory_upgrade": True
        }
        options.add_experimental_option('prefs', prefs)
        options.add_argument(f'--user-data-dir={user_config.user_data_path}')
        options.add_argument(f'--profile-directory={user_config.profile_name}')

        return webdriver.Chrome(chrome_options=options)

    @staticmethod
    def get_chrome_driver_gfa() -> WebDriver:

        options = webdriver.ChromeOptions()
        # https://www.python2.net/questions-80772.htm
        options.add_experimental_option("detach", True)
        user_config = Utils._get_config()
        prefs = {
            "profile.default_content_settings.popups": 0,
            f'download.default_directory': user_config.download_path,
            "directory_upgrade": True
        }
        options.add_experimental_option('prefs', prefs)
        options.add_argument(f'--user-data-dir={user_config.user_data_path}')
        options.add_argument(f'--profile-directory={user_config.gfa_profile_name}')

        return webdriver.Chrome(chrome_options=options)

    @staticmethod
    def _get_config() -> UserConfig:
        # path
        path = RAW_FILE_PATH if RAW_FILE_PATH != "" else f'{os.path.dirname(os.path.abspath(Path(__file__).parent))}{os.sep}raw_data'
        if not os.path.isdir(path):
            os.mkdir(path)

        # user_data_path
        user_data_path = CHROME_USER_DATA_PATH if CHROME_USER_DATA_PATH != "" else "/Users/qualson/Library/Application Support/Google/Chrome/Profile 2"
        profile_name = CHROME_PROFILE_NAME if CHROME_PROFILE_NAME != "" else "Profile 2"
        gfa_profile_name = CHROME_GFA_PROFILE_NAME if CHROME_GFA_PROFILE_NAME != "" else "Profile 3"

        return UserConfig(user_data_path=user_data_path, profile_name=profile_name, gfa_profile_name=gfa_profile_name,
                          download_path=path)

    @staticmethod
    def get_day_name(date: str):

        date_obj = datetime.datetime.strptime(date, '%Y-%m-%d')
        daylist = ["월", "화", "수", "목", "금", "토", "일"]
        return daylist[date_obj.weekday()]

    @staticmethod
    def get_day(ago: float) -> str:
        date = datetime.date.today() - datetime.timedelta(ago)
        return date.isoformat()

    @staticmethod
    def get_today() -> str:
        return datetime.date.today().isoformat()

    @staticmethod
    def get_yesterday() -> str:
        date = datetime.date.today() - datetime.timedelta(1)
        return date.isoformat()

    @staticmethod
    def get_3daysago() -> str:
        date = datetime.date.today() - datetime.timedelta(3)
        return date.isoformat()

    @staticmethod
    def get_weekday() -> int:
        # It's MONDAY when return value is 0
        return datetime.datetime.today().weekday()

    @staticmethod
    def send_keys_delayed(element: WebElement, input: str):
        for char in input:
            element.send_keys(char)
            time.sleep(random.uniform(0.02, 0.1))

    @staticmethod
    def kill_proc(proc_exp: str):
        try:
            for proc in psutil.process_iter():
                if fnmatch.fnmatch(proc.name(), proc_exp):
                    proc.kill()
        except Exception as e:
            print(e)
    @staticmethod
    def create_xl_sheet(wb: Workbook, sheet_name: str) -> worksheet:
        rd_ws_name = sheet_name
        ws = None
        if rd_ws_name in wb.sheetnames:
            ws = wb[rd_ws_name]
        else:
            ws = wb.create_sheet(rd_ws_name)
        return ws

    @staticmethod
    def get_recent_file(expression: str) -> str:
        dir_path = "." + os.sep + "raw_data" + os.sep
        file_path = ""
        ctime = 0
        for file_name in os.listdir(dir_path):
            if fnmatch.fnmatch(file_name, expression):
                if ctime < os.path.getmtime(dir_path + file_name):
                    ctime = os.path.getmtime(dir_path + file_name)
                    file_path = file_name
        return dir_path + file_path

    @staticmethod
    def backup_original_files():
        for domain, file in RD_FILE.items():
            domain_backup_path = f"{RAW_FILE_PATH}/backup/{domain}"
            Path(domain_backup_path).mkdir(parents=True, exist_ok=True)
            copyfile(Utils._get_raw_file_path(file), Utils._get_backup_file_path(domain, file))

    @staticmethod
    def remove_old_backup_files():
        for domain, file in RD_FILE.items():
            domain_backup_path = f"{RAW_FILE_PATH}/backup/{domain}"
            Utils._remove_old_backup_files(domain_backup_path)

    @staticmethod
    def _get_raw_file_path(file: str) -> str:
        return f"{RAW_FILE_PATH}/{file}"

    @staticmethod
    def _get_backup_file_path(domain: str, file: str) -> str:
        now = datetime.datetime.now().strftime("%Y-%m-%d-%H-%M-%S")
        return f"{RAW_FILE_PATH}/backup/{domain}/{now}-{file}"

    @staticmethod
    def _remove_old_backup_files(domain_folder: str):
        files = [join(domain_folder, f) for f in listdir(domain_folder) if isfile(join(domain_folder, f))]
        for file in files:
            modified_time = datetime.datetime.fromtimestamp(Path(file).stat().st_mtime)
            current_time = datetime.datetime.now()
            if current_time.day - modified_time.day > 7:
                os.remove(file)


    @staticmethod
    def vlookup_by_matching(ws: worksheet, matching: str, content: str):
        res = "0"
        content_map = {
            "채널": 2,
            "상품1": 3,
            "상품2": 4,
            "상품 상세": 5,
            "구분(판매가)": 6,
            "구분값": 7,
            "판매가": 8,
            "수수료": 9,
            "원가": 10
            }
        
        max_row = ws.max_row + 1
        for row in range(1, max_row):
            if ws.cell(row = row, column = 4).value == matching:
                res = ws.cell(row = row, column = content_map[content]).value
                break
        return res

    @staticmethod
    def vlookup_by_cutoff(ws: worksheet, cutoff: str, content: str):
        res = "0"
        content_map = {
            "채널": 2,
            "상품1": 3,
            "상품2": 5,
            "구분(판매가)": 6,
            "판매가": 8,
            "수수료": 9,
            "원가": 10
            }
        
        max_row = ws.max_row + 1
        for row in range(1, max_row):
            if ws.cell(row = row, column = 7).value == cutoff:
                res = ws.cell(row = row, column = content_map[content]).value
                break
        return res

    @staticmethod
    def vlookup_ads(ws: worksheet, matching: str, content: str):
        res=""
        content_map = {
            "미디어": 3,
            "상품1": 4,
        }
        max_row = ws.max_row + 1
        for row in range(1, max_row):
            if ws.cell(row = row, column = 2).value == matching:
                res = ws.cell(row = row, column = content_map[content]).value
                break
        return res

    @staticmethod
    def vlookup_cafe24(ws: worksheet, matching: str):
        res = "0"
        max_row = ws.max_row + 1
        for row in range(1, max_row):
            if ws.cell(row = row, column = 2).value == matching:
                res = ws.cell(row = row, column = 3).value
                break
        return res
