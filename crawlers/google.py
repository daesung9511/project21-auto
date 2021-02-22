#coding: utf-8

import time
import io
from googleads import adwords

from openpyxl import load_workbook
from datetime import datetime, timedelta

from utils import Utils, RD_FILE
from config import GOOGLE_ADS_PATH

import math
import csv

class Google:

    def update_ad_fee_data(self, account, datas, workbooks):

        wb = workbooks[account["domain"]]
        
        ws = Utils.create_xl_sheet(wb, "RD")

        for inner_datas in datas:
            if not len(inner_datas) == 0: 
                max_row = str(ws.max_row+1)
                ws.cell(row=int(max_row),column=1).value = (inner_datas[1]).strip("\"")
                ws.cell(row=int(max_row),column=2).value = inner_datas[0]
                ws.cell(row=int(max_row),column=3).value = Utils.get_day_name(inner_datas[0])
                ws.cell(row=int(max_row),column=4).value = Utils.vlookup_ads(wb["매칭테이블"], inner_datas[1], "미디어")
                ws.cell(row=int(max_row),column=5).value = Utils.vlookup_ads(wb["매칭테이블"], inner_datas[1], "상품1")
                ws.cell(row=int(max_row),column=11).value = math.floor(int(inner_datas[2]) / 1000000)

    def get_ad_data(self, account, days):
        
        start_date = (datetime.now() + timedelta(days=-days)).strftime('%Y%m%d')
        end_date =  (datetime.now() + timedelta(days=-1)).strftime('%Y%m%d')
        # Define output as a string
        output = io.StringIO()

        # Initialize appropriate service.
        adwords_client = adwords.AdWordsClient.LoadFromStorage(path=GOOGLE_ADS_PATH)
        report_downloader = adwords_client.GetReportDownloader(version='v201809')

        # Create report query.
        report_query = (f'''
        select Date, CampaignName, Cost
        from CAMPAIGN_PERFORMANCE_REPORT
        during {start_date}, {end_date}''')

        # Write query result to output file
        report_downloader.DownloadReportWithAwql(
            report_query, 
            'CSV',
            output,
            client_customer_id=account["client_customer_id"], # denotes which adw account to pull from
            skip_report_header=True, 
            skip_column_header=True,
            skip_report_summary=True, 
            include_zero_impressions=False)

        output.seek(0)
        datas=[]
        for line in output.readlines():
            inner_datas = []
            a = line.rsplit(",", 1)[0]
            b = line.rsplit(",", 1)[1]
            c = a.split(",", 1)
            inner_datas.extend([c[0], c[1], b])
            datas.append(inner_datas)
        
        return datas
            
    def run(self, driver, account, days, workbooks):
        datas = self.get_ad_data(account, days)
        self.update_ad_fee_data(account, datas, workbooks)
            

