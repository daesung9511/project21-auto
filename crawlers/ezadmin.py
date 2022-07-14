import datetime
import logging
import time

import pandas as pd
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from pandas import DataFrame, Index
from selenium.webdriver.android.webdriver import WebDriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.select import Select
from selenium.webdriver.support.wait import WebDriverWait

from config import CUTOFF_VERSION
from secrets import ACCOUNTS
from utils import Utils, DEFAULT_TIMEOUT_DELAY
import numpy as np

class Ezadmin:

    def run(self, driver, account, days, workbooks):
        uid = account["id"]
        upw = account["pw"]
        udomain = account["udomain"]
        for day in range(days, 0, -1):
            date = (datetime.datetime.now() + datetime.timedelta(days=-day)).strftime('%Y-%m-%d')
            Ezadmin.download_yesterday_revenue(driver, udomain, uid, upw, date)
            print('=================================111111')
            Ezadmin.update_rd_data(account["domain"], date, workbooks)
            print('=================================')

    @staticmethod
    def get_admin_page(driver: WebDriver, domain: str, id: str, password: str) -> WebDriver:
        driver.get("https://www.ezadmin.co.kr/index.html#main")
        WebDriverWait(driver, 3).until(
            expected_conditions.presence_of_all_elements_located((By.ID, "login-popup"))
        )
        driver.find_element_by_css_selector("#header > div.utilWrap > ul > li.login > a").click()
        driver.find_element_by_css_selector("#login-domain").send_keys(domain)
        driver.find_element_by_css_selector("#login-id").send_keys(id)
        pw_element = driver.find_element_by_css_selector("#login-pwd")
        pw_element.clear()
        pw_element.send_keys(password)

        driver.find_element_by_css_selector(
            "#login-popup > div.content-inner > form:nth-child(4) > input.login-btn"
        ).click()

        # Login complete
        try:
            keep_password_selector = "#passwd_keep > a:nth-child(6) > img"
            WebDriverWait(driver, DEFAULT_TIMEOUT_DELAY).until(
                expected_conditions.presence_of_all_elements_located((By.CSS_SELECTOR, keep_password_selector))
            )
            driver.find_element_by_css_selector(keep_password_selector).click()
        except Exception as e:
            logging.warning(e)

        try:
            notice_selector = "#pop_top > span > a > img"
            WebDriverWait(driver, DEFAULT_TIMEOUT_DELAY).until(
                expected_conditions.presence_of_all_elements_located((By.CSS_SELECTOR, notice_selector))
            )
            driver.find_element_by_css_selector(notice_selector).click()
        except Exception as e:
            logging.warning(e)

        return driver

    @staticmethod
    def download_yesterday_revenue(driver: WebDriver, domain: str, id: str, password: str, date: str) -> WebDriver:

        driver = Ezadmin.get_admin_page(driver, domain, id, password)
        time.sleep(12)
        menu_selector = "#mymenu > a"
        driver.find_element_by_css_selector(menu_selector).click()
        time.sleep(2)
        menus = driver.find_elements_by_css_selector("#mymenu_list > li")
        mymenu = None
        for menu in menus:
            a_menu = menu.find_element_by_css_selector('a')
            if a_menu.text == '판매처상품매출통계':
                mymenu = a_menu
        mymenu.click()
        time.sleep(3)
        query_type = '#query_type'
        select = Select(driver.find_element_by_css_selector(query_type))
        select.select_by_value('order_date')

        date_selector = 'input.datepicker.hasDatepicker'
        date_inputs = driver.find_elements_by_css_selector(date_selector)

        for date_input in date_inputs:
            length = len(date_input.get_attribute('value'))
            date_input.send_keys(length * Keys.BACKSPACE)
            date_input.send_keys(date)
            time.sleep(5)

        driver.find_element_by_css_selector('#search').click()
        time.sleep(5)

        download = '/html/body/div[3]/table/tbody/tr/td[3]/div[2]/div[1]/div[3]/div[2]/a/span'
        driver.find_element_by_xpath(download).click()
        WebDriverWait(driver, DEFAULT_TIMEOUT_DELAY).until(
            expected_conditions.alert_is_present()
        )
        driver.switch_to.alert.accept()
        time.sleep(10)

        return driver

    @staticmethod
    def match_columns(columns: Index):

        for column in columns:
            print(column)
            if column == 1 or column == "Unnamed: 1" or column == np.nan or column == "nan":
                return False
        return True

    @staticmethod
    def match_columns_index(columns: Index):
        col_tuple = ["판매처", "판매처 상품명", "판매처상품옵션", "주문수량", "상품수", "상품원가", "판매금액", "판매금액 합", "어드민 상품명", "상품판매가", "상품판매가x수량", "공급처 상품명", "정산금액"]
        for column in columns:
            print(column)
            if column not in col_tuple:
                return False
        return True
    @staticmethod
    def find_the_real_column_name_index(df: DataFrame, count: int = 0):
        if df.shape[0] == count - 1 or len(df) == 0:
            return 0
        if Ezadmin.match_columns_index(df.loc[count]):
            return count
        return Ezadmin.find_the_real_column_name_index(df, count+1)

    @staticmethod
    def update_rd_data(domain: str, date: str, workbooks: dict):

        sales_wb = workbooks[domain]
        # TODO: 해당 날짜 시트겹치는 것 체크
        sales_ws = Utils.create_xl_sheet(sales_wb, "RD")

        # Open Ezadmin rd file
        xls_path = Utils.get_recent_file("판매처상품매출통계*.xls")
        try:
            f = open(xls_path, 'r', encoding='UTF8')
            datas = Ezadmin.parse_html_data(f)
        except Exception as e:
            print('날짜 파싱 오류')
            raise e
        headers = datas.pop(0)
        df_list = pd.read_html(xls_path, header=0)
        df_to_use = df_list[0]
        print(df_to_use)
        print(df_to_use.columns)
        d = {df_to_use.columns[index]: key for index, key in enumerate(["판매처", "판매처 상품명", "판매처상품옵션", "주문수량", "상품수", "상품원가", "판매금액", "판매금액 합", "어드민 상품명", "상품판매가", "상품판매가x수량", "공급처 상품명", "정산금액"])}
        print(f"dictionary: {d}")
        print(df_to_use.columns)
        if not Ezadmin.match_columns(df_to_use.columns):
            print("true")
            real_column_index = Ezadmin.find_the_real_column_name_index(df_to_use)
            df_refresh = df_to_use.rename(columns=d).iloc[real_column_index+1:, :]
        else:
            print("false")
            df_refresh = df_to_use
        print(df_refresh.head())
        for idx, row in df_refresh.fillna({"판매처상품옵션": ""}).iterrows():
            channel = row['판매처']
            matching = row['판매처 상품명'] + row['판매처상품옵션']
            sales = row['주문수량']
            if Utils.exclude_by_keyword(domain, channel):
                continue
            sales_max_row = str(sales_ws.max_row + 1)
            prod1 = Utils.vlookup_by_matching(sales_wb["매칭테이블"], matching, "상품1")

            cur_cutoff = CUTOFF_VERSION[domain]
            cutoff = channel + prod1 + matching + cur_cutoff
            try:
                sales_ws["B" + sales_max_row].value = datetime.datetime.strptime(date, '%Y-%m-%d').date()
                sales_ws["C" + sales_max_row].value = Utils.get_day_name(date)
                sales_ws["E" + sales_max_row].value = prod1
                sales_ws["F" + sales_max_row].value = channel
                sales_ws["G" + sales_max_row].value = matching
                sales_ws["H" + sales_max_row].value = sales
                sales_ws["I" + sales_max_row].value = Utils.vlookup_by_matching(sales_wb["매칭테이블"], matching, "상품 상세")
                sales_ws["J" + sales_max_row].value = cur_cutoff
                sales_ws["L" + sales_max_row].value = float(int(
                    Utils.vlookup_by_cutoff(sales_wb["매칭테이블"], cutoff, "판매가"))) * float(sales)
                sales_ws["M" + sales_max_row].value = (1 - float(
                    Utils.vlookup_by_cutoff(sales_wb["매칭테이블"], cutoff, "수수료"))) * float(
                    Utils.vlookup_by_cutoff(sales_wb["매칭테이블"], cutoff, "판매가")) * float(sales)
                sales_ws["N" + sales_max_row].value = float(int(
                    Utils.vlookup_by_cutoff(sales_wb["매칭테이블"], cutoff, "원가"))) * float(sales)
                sales_ws["O" + sales_max_row].value = cutoff
            except Exception as e:
                print("sheet line : " + sales_max_row)
                print(e)

    @staticmethod
    def wait(self, driver, selector, sec):
        try:
            WebDriverWait(driver, sec).until(
                expected_conditions.presence_of_element_located((By.CSS_SELECTOR, selector))
            )
            return True
        except Exception as e:
            return False

    @staticmethod
    def parse_html_data(input):

        soup = BeautifulSoup(input, 'html.parser')
        lines = soup.select("tr")
        data = []
        for line in lines:
            # print(row)
            cols = line.find_all('td')
            cols = [ele.text for ele in cols]
            data.append(cols)
        return data


if __name__ == '__main__':
    account = ACCOUNTS["ezadmin"]['project21']
    days = 1
    uid = account["id"]
    upw = account["pw"]
    # driver = Utils.get_chrome_driver()
    # driver.set_window_size(1980, 1080)
    udomain = account["udomain"]
    workbooks = {
        'project21': load_workbook('./02.프로젝트21_이지어드민_데이터 정리_실데이터_헤더변경 ★.xlsx')
    }
    for day in range(days, 0, -1):
        date = (datetime.datetime.now() + datetime.timedelta(days=-day)).strftime('%Y-%m-%d')
        # Ezadmin.download_yesterday_revenue(driver, udomain, uid, upw, date)
        Ezadmin.update_rd_data(account["domain"], date, workbooks)
