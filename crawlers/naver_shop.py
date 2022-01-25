# coding: utf-8

import csv
import json
import time
import datetime

from powernad.Object.AdGroup.AdgroupObject import AdgroupObject
from powernad.Object.Campaign.CampaignObject import CampaignObject
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions

from utils import Utils, DEFAULT_TIMEOUT_DELAY, RD_FILE

from openpyxl import load_workbook

from utils.naverShop import AsCampaign, AsStat


class Naver_shop:

    @staticmethod
    def get_n_days_past_data(timedelta_days, account):
        id = account["account_id"]
        license = account["license"]
        secret = account["secret"]
        
        campaign = AsCampaign("https://api.naver.com", license, secret, id)
        c_list: list[CampaignObject] = campaign.get_campaign_list_with_customer_id(id)
        c_ids = list(map(lambda x: x.nccCampaignId, c_list))
        
        info = []
        for c_id in c_ids:
            campaign = AsCampaign("https://api.naver.com", license, secret, id)
            ad_groups: list[AdgroupObject] = campaign.get_adgroup_list(id, c_id)
            ad_group_ids = list(map(lambda x: x.nccAdgroupId, ad_groups))
            if not ad_group_ids:
                continue

            stat = AsStat("https://api.naver.com", license, secret, id)
            current = datetime.datetime.now() - datetime.timedelta(days=timedelta_days)

            date_format = "%Y-%m-%d"

            current_date = current.strftime(date_format)

            # fields = """[\"clkCnt\",\"impCnt\",\"salesAmt\",\"ctr\",\"cpc\",\"ccnt\",\"crto\",\"convAmt\",\"ror\",\"cpConv\",\"viewCnt\"]"""
            fields = """[\"salesAmt\"]"""
            range = """{\"since\":\"""" + current_date + """\",\"until\":\"""" + current_date + """\"}"""


            # Check if ads_ids empty
            stats = stat.get_stat_by_ids(
                ids=ad_group_ids,
                fields=fields,
                timeRange=range,
                timeIncrement="allDays"
            )

            try:
                sales_info_list = stats['data']
                for ad_group in ad_groups:
                    for x in sales_info_list:
                        sales_info = dict(x)
                        if sales_info["id"] == ad_group.nccAdgroupId:
                            result = dict({"id": sales_info["id"], "name": ad_group.name, "cost": sales_info["salesAmt"], "date": current_date})
                            info.append(result)
                            break
            except Exception as e:
                print(e)
                raise e
        return info

    def get_data(self, account, days):
        for n in range(1, days + 1):
            info = self.get_n_days_past_data(n, account)
            print(info)

    @staticmethod
    def update_ad_costs(domain, datas, workbooks):
        # RD 엑셀 파일 로딩
        wb = workbooks[domain]
        ws = Utils.create_xl_sheet(wb, "RD")
        for data in datas:
            max_row = str(ws.max_row+1)
            
            ws.cell(row=int(max_row),column=1).value = data["name"]
            ws.cell(row=int(max_row),column=2).value = datetime.datetime.strptime(data["date"], '%Y-%m-%d').date()
            ws.cell(row=int(max_row),column=3).value = Utils.get_day_name(data["date"])
            ws.cell(row=int(max_row),column=4).value = Utils.vlookup_ads(wb["매칭테이블"], data["name"], "미디어")
            ws.cell(row=int(max_row),column=5).value = Utils.vlookup_ads(wb["매칭테이블"], data["name"], "상품1")
            ws.cell(row=int(max_row),column=11).value = float(data["cost"])/1.1

        wb.save(RD_FILE[domain])

    def run(self, driver, account, days, workbooks):
        # account list
        # lavena, yuge, anua, project21
        
        for n in range(1, days + 1):
            datas = self.get_n_days_past_data(n, account)
            self.update_ad_costs(account["domain"], datas, workbooks)


if __name__ == '__main__':
    account = {
            "id": "pista1004",
            "pw": "pista1004!",
            "type": "general",
            "account_id": 1154389,
            "license": "0100000000279e9b62c03c168ab8fcd16caa07dbf84f4c7d29e781db4005ca8677611f5f1c",
            "secret": "AQAAAAAnnptiwDwWirj80WyqB9v4C0G9m6Fw5kcRPdPTcrvU+w==",
            "domain": "project21"
        }
    workbooks = {
        'project21': load_workbook('./../02.프로젝트21_이지어드민_데이터 정리_실데이터_헤더변경 ★.xlsx')
    }
    datas = Naver_shop.get_n_days_past_data(1, account)
    for data in datas:
        print(data)
    Naver_shop.update_ad_costs(account["domain"], datas, workbooks)
    workbooks['project21'].save('./../02.프로젝트21_이지어드민_데이터 정리_실데이터_헤더변경 ★.xlsx')

    

