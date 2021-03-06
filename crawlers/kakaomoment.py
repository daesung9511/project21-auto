# coding: utf-8

import csv
import datetime
import time

from openpyxl import load_workbook
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.ui import WebDriverWait

from secrets import ACCOUNTS
from utils import Utils, DEFAULT_TIMEOUT_DELAY


class Kakaomoment:
    flag = True

    @staticmethod
    def switch_popup(driver):
        windows = driver.window_handles
        driver.switch_to.window(windows[-1])

    @staticmethod
    def switch_main(driver):
        windows = driver.window_handles
        driver.switch_to.window(windows[0])

    @staticmethod
    def wait(driver, selector, sec):
        try:
            WebDriverWait(driver, sec).until(
                expected_conditions.presence_of_element_located((By.CSS_SELECTOR, selector))
            )
        except:
            pass

    def wait_popup(self, driver, selector, sec):
        while True:
            try:
                self.switch_popup(driver)
                WebDriverWait(driver, sec).until(
                    expected_conditions.presence_of_element_located((By.CSS_SELECTOR, selector))
                )
                break
            except:
                pass

    @staticmethod
    def close_popup(driver):
        windows = driver.window_handles
        main = windows[0]
        for window in windows:
            if window != main:
                driver.switch_to.window(window)
                driver.close()

        driver.switch_to.window(main)

    def init(self, driver, url):
        driver.get(url)

        # Wait for browser loading
        kakao_login_button = """#login-form > fieldset > div.wrap_btn > button.btn_g.btn_confirm.submit"""
        self.wait(driver, kakao_login_button, 10)

        return driver

    def login(self, driver, account):
        time.sleep(10)
        # get kakao id form
        id_form = driver.find_element_by_id("id_email_2")
        id_form.send_keys(account["id"])

        # get kakao pw form
        pw_form = driver.find_element_by_id("id_password_3")
        pw_form.send_keys(account["pw"])

        # get kakao login button
        kakao_login_button = driver.find_element_by_css_selector(
            """#login-form > fieldset > div.wrap_btn > button.btn_g.btn_confirm.submit""")
        kakao_login_button.click()
        time.sleep(10)
        # wait for login
        dashboard = """#kakaoContent > div.cont_feature > ul > li.on > a"""

        try:
            self.wait(driver, dashboard, DEFAULT_TIMEOUT_DELAY)
        except Exception as e:
            time.sleep(1)

    def move_dashboard_anua(self, driver, number):
        # dashboard url 
        dashboard_url = f"https://moment.kakao.com/{number}/report/customreport/all"

        # move to dashboard
        driver.get(dashboard_url)

        # find report name
        report_name = """#mArticle > div > div.ad_managebox > div.tblg2_wrap > table > tbody > tr:nth-child(1) > td:nth-child(2) > div > a"""
        self.wait(driver, report_name, DEFAULT_TIMEOUT_DELAY)

        pattern = "?????????"
        try:
            for i in range(1, 10 + 1):
                report_name = f"""#mArticle > div > div.ad_managebox > div.tblg2_wrap > table > tbody > tr:nth-child({i}) > td:nth-child(2) > div > a"""
                report_name_element = driver.find_element_by_css_selector(report_name)
                if pattern == report_name_element.text:
                    report_name_element.click()
                    break
        except Exception as e:
            print(e)

        date_form = """#mArticle > div > div.set_table > div.set_head > div.f_right > div:nth-child(3) > div > div.btn_gm.gm_calendar > a"""
        self.wait(driver, date_form, DEFAULT_TIMEOUT_DELAY)

    def move_dashboard(self, driver, number):
        # dashboard url 
        dashboard_url = f"https://moment.kakao.com/dashboard/{number}"

        # move to dashboard
        driver.get(dashboard_url)

        date_form = """#mArticle > div > div.set_table > div.set_head > div.f_right > div:nth-child(2) > div > a"""
        self.wait(driver, date_form, DEFAULT_TIMEOUT_DELAY)

    @staticmethod
    def calc_date(day):
        today = datetime.date.today()
        token = datetime.timedelta(day)

        day_obj = today - token

        # prev month
        if day_obj.day > today.day:
            stat = 1
        else:
            stat = 0

        return stat, day_obj

    def select_date(self, driver, domain, day):
        if domain == "anua":
            date_form = """#mArticle > div > div.set_table > div.set_head > div.f_right > div:nth-child(3) > div > div.btn_gm.gm_calendar > a"""
            yesterday_button = """#mArticle > div > div.set_table > div.set_head > div.f_right > div:nth-child(3) > div > div.btn_gm.gm_calendar.open > div > div > div.layer_body > ul > li.on > a"""
            ok_button = """#mArticle > div > div.set_table > div.set_head > div.f_right > div:nth-child(3) > div > div.btn_gm.gm_calendar.open > div > div > div.layer_body > div > div.btn_wrap > button.btn_gm.gm_bl > span"""
        else:
            date_form = """#mArticle > div > div.adboardbox_search > div.inner_g > div.wrap_calendar > div > div.btn_gm.gm_calendar > a"""
            yesterday_button = """#mArticle > div > div.adboardbox_search > div.inner_g > div.wrap_calendar > div > div.btn_gm.gm_calendar.open > div > div > div.layer_body > ul > li.on > a"""
            ok_button = """#mArticle > div > div.adboardbox_search > div.inner_g > div.wrap_calendar > div > div.btn_gm.gm_calendar.open > div > div > div.layer_body > div > div.btn_wrap > button.btn_gm.gm_bl > span"""

            # date_form = """#mArticle > div > div.dashboard_check > div.f_right > div:nth-child(1) > div > div.btn_gm.gm_calendar > a"""
            # yesterday_button = """#mArticle > div > div.dashboard_check > div.f_right > div:nth-child(1) > div > div.btn_gm.gm_calendar.open > div > div > div.layer_body > ul > li:nth-child(2) > a"""
            # ok_button = """#mArticle > div > div.dashboard_check > div.f_right > div:nth-child(1) > div > div.btn_gm.gm_calendar.open > div > div > div.layer_body > div > div.btn_wrap > button.btn_gm.gm_bl > span"""

        # click date form
        self.wait(driver, date_form, DEFAULT_TIMEOUT_DELAY)
        date_form = driver.find_element_by_css_selector(date_form)
        date_form.click()

        self.wait(driver, yesterday_button, DEFAULT_TIMEOUT_DELAY)

        stat, day_obj = self.calc_date(day)

        # get calendar elements
        calendar = driver.find_elements_by_css_selector("div.area_calendar")
        left = calendar[0]
        right = calendar[1]

        days = ".inner_link_day"
        prev_days = left.find_elements_by_css_selector(days)
        current_days = right.find_elements_by_css_selector(days)

        # stat 0 : start-right, end-right
        if stat == 0:
            for current_day in current_days:
                if current_day.text == str(day_obj.day):
                    current_day.click()
                    current_day.click()
                    break

            driver.implicitly_wait(1)

        # stat 1 : start-left, end-left
        elif stat == 1:
            for prev_day in prev_days:
                if prev_day.text == str(day_obj.day):
                    prev_day.click()
                    prev_day.click()
                    break

            driver.implicitly_wait(1)

        # click ok button
        self.wait(driver, ok_button, DEFAULT_TIMEOUT_DELAY)
        ok_button = driver.find_element_by_css_selector(ok_button)
        ok_button.click()

        # wait for loading
        time.sleep(2)

    def download_csv(self, driver, domain):
        if domain == "anua":
            download_button = """#mArticle > div > div.set_table > div.set_head > div.f_right > div:nth-child(4) > a > span > span"""
        else:
            download_button = """#mArticle > div > div.set_table > div.set_head > div.f_right > div:nth-child(2) > a > span"""

        self.wait(driver, download_button, DEFAULT_TIMEOUT_DELAY)
        download_button = driver.find_element_by_css_selector(download_button)
        download_button.click()

        driver.implicitly_wait(1)
        time.sleep(2)

    @staticmethod
    def logout(driver):
        driver.get(
            "https://accounts.kakao.com/logout?continue=https://accounts.kakao.com/login/kakaoforbusiness?continue=https://business.kakao.com/dashboard/?sid=kmo&redirect=https://moment.kakao.com/dashboard")

    @staticmethod
    def update_ad_costs(account, day, workbooks):
        domain = account["domain"]
        filename = account["filename"]

        # RD ?????? ?????? ??????
        wb = workbooks[domain]
        ws = Utils.create_xl_sheet(wb, "RD")

        date = (datetime.datetime.now() + datetime.timedelta(days=-day)).strftime('%Y-%m-%d')
        if domain == "anua":

            # ???????????????????????? ?????? ?????? ?????? csv ?????? ??????
            anua_path = Utils.get_recent_file(filename)

            with open(anua_path, 'r', encoding='utf-16') as f:
                reader = csv.reader(f, delimiter="\t")
                next(reader)
                # ????????? ????????? rd ??????
                for row in reader:

                    if float(row[4]) == 0:
                        continue

                    max_row = str(ws.max_row + 1)

                    ws.cell(row=int(max_row), column=1).value = row[1]
                    ws.cell(row=int(max_row), column=2).value = datetime.datetime.strptime(row[2], '%Y-%m-%d').date()
                    ws.cell(row=int(max_row), column=3).value = Utils.get_day_name(row[2])
                    ws.cell(row=int(max_row), column=4).value = Utils.vlookup_ads(wb["???????????????"], row[1], "?????????")
                    ws.cell(row=int(max_row), column=5).value = Utils.vlookup_ads(wb["???????????????"], row[1], "??????1")
                    ws.cell(row=int(max_row), column=11).value = float(row[4]) / 1.1

        else:
            # ???????????? csv ??????
            file_path = Utils.get_recent_file(filename)
            with open(file_path, 'r', encoding='utf-16') as f:
                reader = csv.reader(f, delimiter="\t")
                next(reader)
                # ????????? ????????? rd ??????
                for row in reader:
                    if float(row[6]) == 0:
                        continue
                    # # "?????? ???" ????????? ???????????? ??????
                    # if not row[2] == "?????? ???":
                    #     continue

                    max_row = str(ws.max_row + 1)

                    ws.cell(row=int(max_row), column=1).value = row[0]
                    ws.cell(row=int(max_row), column=2).value = datetime.datetime.strptime(date, '%Y-%m-%d').date()
                    ws.cell(row=int(max_row), column=3).value = Utils.get_day_name(date)
                    ws.cell(row=int(max_row), column=4).value = Utils.vlookup_ads(wb["???????????????"], row[0], "?????????")
                    ws.cell(row=int(max_row), column=5).value = Utils.vlookup_ads(wb["???????????????"], row[0], "??????1")
                    ws.cell(row=int(max_row), column=11).value = float(row[6]) / 1.1

    def run(self, driver, account, term, workbooks):
        # account list
        # lavena, yuge, anua, project21

        Utils.kill_proc("chrome*")
        driver = Utils.get_chrome_driver_gfa()
        driver.set_window_size(1980, 1080)

        url = "https://accounts.kakao.com/login/kakaoforbusiness?continue=https://business.kakao.com/dashboard/?sid=kmo&redirect=https://moment.kakao.com/dashboard"

        if self.flag:
            self.init(driver, url)
            self.login(driver, account)

        for day in range(term, 0, -1):
            if account["domain"] == "anua":
                self.move_dashboard_anua(driver, account["number"])
            else:
                self.move_dashboard(driver, account["number"])
            self.select_date(driver, account["domain"], day)
            self.download_csv(driver, account["domain"])
            self.update_ad_costs(account, day, workbooks)

        self.flag = False
