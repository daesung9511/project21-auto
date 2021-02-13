# coding: utf-8

import datetime
import time

from selenium.webdriver.chrome.webdriver import WebDriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.ui import WebDriverWait

from openpyxl import load_workbook
import os
import fnmatch
import csv

from utils import Utils, DEFAULT_TIMEOUT_DELAY, RD_FILE

class Naver_GFA:
    flag = True

    def switch_popup(self, driver):
        windows = driver.window_handles
        driver.switch_to.window(windows[-1])

    def switch_main(self, driver):
        windows = driver.window_handles
        driver.switch_to.window(windows[0])

    def wait(self, driver, selector, sec):
        try:
            WebDriverWait(driver, sec).until(
                expected_conditions.presence_of_element_located((By.CSS_SELECTOR, selector))
            )
            return True
        except Exception as e:
            return False

    def wait_xpath(self, driver, selector, sec):
        try:
            WebDriverWait(driver, sec).until(
                expected_conditions.presence_of_element_located((By.XPATH, selector))
            )
            return True
        except Exception as e:
            return False

    def wait_popup(self, driver, selector, sec):
        while True:
            try:
                self.switch_popup(driver)
                WebDriverWait(driver, sec).until(
                    expected_conditions.presence_of_element_located((By.CSS_SELECTOR, selector))
                )
                break
            except Exception as e:
                print(e)

    def close_popup(self, driver):
        windows = driver.window_handles
        main = windows[0]
        for window in windows:
            if window != main:
                driver.switch_to.window(window)
                driver.close()

        driver.switch_to.window(main)

    def init(self, driver, url):
        driver.get(url)

        # Wait for browser loading
        naver_login_button = """body > div > div.container.bg_white > div > div > div.login_box > ul > li.selected > a"""
        self.wait(driver, naver_login_button, 10)

        return driver

    def login(self, driver, account):
        # get naver login button
        naver_login_button = driver.find_element_by_css_selector(
            """body > div > div.container.bg_white > div > div > div.login_box > ul > li.selected > a > span.platform_name""")
        naver_login_button.click()

        # wait for login popup
        naver_banner = """#log\.login"""
        self.wait(driver, naver_banner, DEFAULT_TIMEOUT_DELAY)

        # get naver id form
        id_form = driver.find_element_by_id("id")
        Utils.send_keys_delayed(id_form, account["id"])

        # get naver pw form
        pw_form = driver.find_element_by_id("pw")
        Utils.send_keys_delayed(pw_form, account["pw"])

        # get login button
        login_button = driver.find_element_by_id("log.login")
        login_button.click()

    def press_ok(self, driver):
        # wait for ok button after login
        login_ok_button = """#app > div > div.container > div.content > div > div.panel_body > span > div > div > div.ly_content > div.ly_footer.type_border > button"""

        try:
            self.wait(driver, login_ok_button, 2)
            driver.find_element_by_css_selector(login_ok_button).click()
        except:
            pass

    def select_date(self, driver, account, day):

        date = datetime.date.today() - datetime.timedelta(day)

        query_url = "https://gfa.naver.com/adAccount/accounts/" + account["account_id"] + \
                "/report/performance?dateFirst=" + str(date) + "&dateSecond=" + str(date) + \
                "&adUnit=CAMPAIGN&dateUnit=DAY&placeUnit=TOTAL&dimension=TOTAL&filterList=%5B%5D&showColList=%5B%22col_result%22,%22col_sales_per_result%22,%22col_sales%22,%22col_schedule%22,%22col_imp_count%22,%22col_cpm%22,%22col_click_count%22,%22col_cpc%22,%22col_ctr%22%5D" + \
                "&currentPage=1&pageSize=10&accessAdAccountNo=" + account["account_id"]
        
        driver.get(query_url)

    def download_csv(self, driver, domain):
        if domain == "anua":
            download_button = "#app > div > div.container > div.content > div > div.panel_body.report > div:nth-child(2) > div > div > div.ad_title > div > div.inner_right > a > button"
        else:
            download_button = "#app > div > div.container > div.content > div > div.panel_body > div:nth-child(2) > div > div > div.ad_title > div > div.inner_right > a > button"
        
        self.wait(driver, download_button, 10)
        driver.find_element_by_css_selector(download_button).click()

        driver.implicitly_wait(1)
        time.sleep(2)

    def logout(self, driver):
        logout_button = "li.logout > a"
        driver.find_element_by_css_selector(logout_button).click()

    def clear_tabs(self, driver):
        windows = driver.window_handles
        main = windows[-1]
        for window in windows:
            if window != main:
                driver.switch_to.window(window)
                driver.close()

        driver.switch_to.window(main)

    def update_ad_costs(self, domain, day, workbooks):
    
        # RD 엑셀 파일 로딩
        wb = workbooks[domain]

        ws = Utils.create_xl_sheet(wb, "RD")

        date = (datetime.datetime.now() + datetime.timedelta(days=-day)).strftime('%Y-%m-%d')

        if domain == "yuge":
            file_path = Utils.get_recent_file("유즈_성과리포트_*.csv")
        elif domain == "anua":
            file_path = Utils.get_recent_file("더파운더즈_성과리포트_*.csv")
        elif domain == "lavena":
            file_path = Utils.get_recent_file("라베나코리아_성과리포트_*.csv")

        with open(file_path, 'r', encoding='utf-8') as f:
            reader = csv.reader(f, delimiter = ",")
            next(reader)
            # 광고비 시트에 rd 대입
            for row in reader:
                
                fee_max_row = str(ws.max_row+1)
                
                ws.cell(row=int(fee_max_row),column=1).value = row[0]
                ws.cell(row=int(fee_max_row),column=2).value = date
                ws.cell(row=int(fee_max_row),column=3).value = Utils.get_day_name(date)
                ws.cell(row=int(fee_max_row),column=4).value = Utils.vlookup_ads(wb["매칭테이블"], row[0], "미디어")
                ws.cell(row=int(fee_max_row),column=5).value = Utils.vlookup_ads(wb["매칭테이블"], row[0], "상품1")
                ws.cell(row=int(fee_max_row),column=11).value = float(row[12])/1.1


    def run(self, driver, account, term, workbooks):
        # account list
        # lavena, yuge, anua, project21

        url = "https://auth.glad.naver.com/login?destination=http://gfa.naver.com/adAccount"
        
        Utils.kill_proc("chrome*")
        driver = Utils.get_chrome_driver_gfa()
        driver.set_window_size(1980, 1080)
        
        self.init(driver, url)
        naver_login_button = driver.find_element_by_css_selector(
            """body > div > div.container.bg_white > div > div > div.login_box > ul > li.selected > a > span.platform_name""")
        
        naver_login_button.click()
        time.sleep(3)
        self.press_ok(driver)
        for day in range(term, 0, -1):
            self.select_date(driver, account, day)
            self.download_csv(driver, account["domain"])
            self.clear_tabs(driver)
            self.update_ad_costs(account["domain"], day, workbooks)
        
        self.flag = False