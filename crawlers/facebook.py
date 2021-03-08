from facebook_business.adobjects.campaign import Campaign
from facebook_business.adobjects.adsinsights import AdsInsights
from facebook_business.api import FacebookAdsApi
from facebook_business.adobjects.adaccount import AdAccount
from facebook_business.adobjects.adaccountuser import AdAccountUser

from openpyxl import load_workbook
from datetime import datetime, timedelta

from utils import Utils, RD_FILE
from secrets import ACCOUNTS

class Facebook:
    def update_ad_fee_data(self, account, term, workbooks):

        domain = account["domain"]

        app = ACCOUNTS["facebook"]["app"]

        FacebookAdsApi.init(app["id"], app["secret"], app["access_token"])
        
        wb = workbooks[domain]
        
        ws = Utils.create_xl_sheet(wb, "RD")
        
        campaigns = list(AdAccount(account["id"]).get_campaigns(
        fields=[
                'id',
                'name'
            ]   ,
        params={
                'effective_status': ['ACTIVE'],
            },
        ))

        for day in range(term, 0, -1):
            
            date = (datetime.now() + timedelta(days=-day)).strftime('%Y-%m-%d')
            time_range = { 
                    'since' : date,
                    'until' : date
                }
            
            for campaign in campaigns:
                insights = Campaign(campaign["id"]).get_insights(
                    params={
                        'time_range': time_range
                    }
                )
                
                if not len(insights) == 0:
                    insight = insights[0] 
                    max_row = str(ws.max_row+1)
                    ws.cell(row=int(max_row),column=1).value = campaign["name"]
                    ws.cell(row=int(max_row),column=2).value = datetime.datetime.strptime(date, '%Y-%m-%d').date()
                    ws.cell(row=int(max_row),column=3).value = Utils.get_day_name(date)
                    ws.cell(row=int(max_row),column=4).value = Utils.vlookup_ads(wb["매칭테이블"], campaign["name"], "미디어")
                    ws.cell(row=int(max_row),column=5).value = Utils.vlookup_ads(wb["매칭테이블"], campaign["name"], "상품1")
                    ws.cell(row=int(max_row),column=11).value = float(insight["spend"])
 
    def run(self, driver, account, term, workbooks):
        self.update_ad_fee_data(account, term, workbooks)

