from selenium.webdriver.android.webdriver import WebDriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.wait import WebDriverWait
import logging

from openpyxl.styles import PatternFill, Color
from openpyxl import Workbook
from openpyxl import load_workbook

import time

import datetime
import csv
import os
import fnmatch


from utils import Utils, DEFAULT_TIMEOUT_DELAY, RD_FILE


class Cafe24:
    def run(self, driver, account, days, workbooks):
        uid = account["id"]
        upw = account["pw"]
        products = ["유산균", "하루채움"]
        for day in range(days, 0, -1):
            for product in products:
                Cafe24.download_lacto_revenue(driver, uid, upw, day, product)
                Cafe24.update_rd_data("project21", day, workbooks, product)

    @staticmethod
    def get_admin_page(driver: WebDriver, id: str, password: str) -> WebDriver:
        
        driver.get("https://eclogin.cafe24.com/Shop/")
        id_selector = "#mall_id"
        WebDriverWait(driver, 5).until(
            expected_conditions.presence_of_all_elements_located((By.CSS_SELECTOR, id_selector))
        )
        driver.find_element_by_css_selector(id_selector).send_keys(id)
        driver.find_element_by_css_selector("#userpasswd").send_keys(password)
        driver.find_element_by_css_selector(
            "#frm_user > div.tabCont > div.mButton > button").click()

        # Login complete
        return driver

    @staticmethod
    def download_lacto_revenue(driver:WebDriver, id: str, password: str, day: float, product: str) -> WebDriver:
        
        driver = Cafe24.get_admin_page(driver, id, password)
        time.sleep(5)
        driver.get("https://project21.cafe24.com/disp/admin/shop1/report/ProductPrdchart")

        report_base_url = "https://project21.cafe24.com/disp/admin/shop1/report/ProductPrdchart"
        product_name = product
        start_date = Utils.get_day(day)
        end_date = Utils.get_day(day)
        report_query = f"?searchDateRange=7&eOrderProductCode=&eOrderProductNo=&eOrderProductTag=&rows=10&sOrderBy=sale_cnt&sType=item&excel_public_auth=T&start_date={start_date}&end_date={end_date}&sCategory-1=&eProductSearchType=product_name&eOrderProductText={product_name}&sManufacturerCode=&sTrendCode=&sBrandCode=&sSupplierCode=&sClassificationCode=&iStartProductPrice=0&iEndProductPrice=0&sMobileFlag=ALL&sOverseaFlag=ALL"

        driver.get(f"{report_base_url}{report_query}")

        original_window = driver.current_window_handle

        request_excel_selector = "#QA_product3 > div.mState > div.gRight > div > a > span"
        WebDriverWait(driver, DEFAULT_TIMEOUT_DELAY).until(
            expected_conditions.presence_of_element_located((By.CSS_SELECTOR, request_excel_selector))
        )
        driver.find_element_by_css_selector(request_excel_selector).click()

        excel_download_selector = "#eExcelDownloadButton"
        WebDriverWait(driver, DEFAULT_TIMEOUT_DELAY).until(
            expected_conditions.presence_of_element_located((By.CSS_SELECTOR, excel_download_selector))
        )
        driver.find_element_by_css_selector(excel_download_selector).click()

        WebDriverWait(driver, DEFAULT_TIMEOUT_DELAY).until(
            expected_conditions.alert_is_present()
        )
        driver.switch_to.alert.dismiss()

        download_done = False

        while not download_done:

            # open download popup
            driver.find_element_by_css_selector("#eExcelDownloadPopup > span").click()

            for handle in driver.window_handles:
                driver.switch_to.window(handle)

            first_date = driver.find_element_by_css_selector("#eFileListBody > tr:nth-child(1) > td:nth-child(1)").text
            if first_date == f"{start_date} ~ {end_date}":
                first_date_selector = "#eFileListBody > tr:nth-child(1) > td:nth-child(4)"
                if driver.find_element_by_css_selector(first_date_selector).text == "엑셀다운로드":
                    driver.find_element_by_css_selector(
                        "#eFileListBody > tr:nth-child(1) > td:nth-child(4) > a > span").click()
                    download_done = True
                else:
                    logging.debug("File is not ready retry download")
                    driver.implicitly_wait(1)  # wait for some delay

                driver.close()
            else:
                logging.debug("First excel sheet is not correct")
                # Stop download loop because something wrong happend
                download_done = True
            driver.switch_to.window(original_window)

        return driver

    @staticmethod
    def update_rd_data(domain: str, day: float, workbooks: dict, product: str):
        
        # 매칭테이블 엑셀 파일 로딩 (sales 매칭테이블))
        sales_wb = workbooks[domain]

        # TODO: 해당 날짜 시트겹치는 것 체크
        sales_ws = Utils.create_xl_sheet(sales_wb, "RD")

        cafe24_ws = Utils.create_xl_sheet(sales_wb, "카페24 RD")

        # Cafe24에서 받은 csv 파일 찾기
        csv_path = Utils.get_recent_file("*_ProductPrdchart.csv") 

        with open(csv_path, 'r', encoding='UTF8') as f:
            reader = csv.reader(f)
            next(reader)    # 첫행(헤더 셀) 무시

            date = (datetime.datetime.now() + datetime.timedelta(days=-day)).strftime('%Y-%m-%d')
            # 카페24 RD 피벗테이블 작성
            if product == "하루채움":
                for row in reader:
                    
                    cafe24_max_row = cafe24_ws.max_row + 1
                    sales_max_row = str(sales_ws.max_row+1)
                    matching = row[2] + row[3]
                    prod2 = Utils.vlookup_cafe24(sales_wb["카페24 매칭"], matching)

                    for idx, data in enumerate(row):
                        cafe24_ws.cell(row=cafe24_max_row, column=1).value = date
                        cafe24_ws.cell(row=cafe24_max_row, column=2).value = matching
                        cafe24_ws.cell(row=cafe24_max_row, column=3).value = prod2
                        cafe24_ws.cell(row=cafe24_max_row, column=idx + 4).value = data

            elif product == "유산균":
                dict = {}
                for row in reader:
                    
                    cafe24_max_row = cafe24_ws.max_row + 1
                    matching = row[2] + row[3]
                    prod2 = Utils.vlookup_cafe24(sales_wb["카페24 매칭"], matching)
                    sales=int(row[8].replace(",",""))

                    if prod2 == "0":
                        # Remove suffix of matching
                        matching = matching.split("-", -1)[0]
                        prod2 = Utils.vlookup_cafe24(sales_wb["카페24 매칭"], matching)
                        
                    if prod2 in dict:
                        dict[prod2] += sales
                    else:
                        dict[prod2] = sales

                    for idx, data in enumerate(row):
                        cafe24_ws.cell(row=cafe24_max_row, column=1).value = date
                        cafe24_ws.cell(row=cafe24_max_row, column=2).value = matching
                        cafe24_ws.cell(row=cafe24_max_row, column=3).value = prod2
                        cafe24_ws.cell(row=cafe24_max_row, column=idx + 4).value = data

                for matching in dict.keys():
                    sales_max_row = str(sales_ws.max_row+1)
                    prod1 = Utils.vlookup_by_matching(sales_wb["매칭테이블"], matching, "상품1")
                    channel = Utils.vlookup_by_matching(sales_wb["매칭테이블"], matching, "채널")
                    sales = dict[matching]
                    # TODO: 구분 (ex. 210201) 값이 변동할시 어떻게 적용할지
                    cur_cutoff = "210201"
                    cutoff = channel+prod1+matching+cur_cutoff

                    sales_ws["B" + sales_max_row].value = date
                    sales_ws["C" + sales_max_row].value = Utils.get_day_name(date)
                    sales_ws["E" + sales_max_row].value = prod1
                    sales_ws["F" + sales_max_row].value = channel
                    sales_ws["G" + sales_max_row].value = matching
                    sales_ws["H" + sales_max_row].value = sales
                    sales_ws["I" + sales_max_row].value = Utils.vlookup_by_matching(sales_wb["매칭테이블"], matching, "상품 상세")
                    sales_ws["J" + sales_max_row].value = cur_cutoff
                    sales_ws["L" + sales_max_row].value = int(Utils.vlookup_by_cutoff(sales_wb["매칭테이블"], cutoff, "판매가")) * sales
                    sales_ws["M" + sales_max_row].value = (100.0-float(Utils.vlookup_by_cutoff(sales_wb["매칭테이블"], cutoff, "수수료").strip("%"))) / 100.0 * float(Utils.vlookup_by_cutoff(sales_wb["매칭테이블"], cutoff, "판매가")) * sales
                    sales_ws["N" + sales_max_row].value = int(Utils.vlookup_by_cutoff(sales_wb["매칭테이블"], cutoff, "원가")) * sales
                    sales_ws["O" + sales_max_row].value = int(sales_ws["L" + sales_max_row].value) / 1.1
                    sales_ws["P" + sales_max_row].value = cutoff
